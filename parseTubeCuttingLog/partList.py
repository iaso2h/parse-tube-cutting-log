import config
import console
import util

import re
import os
import datetime
import win32api, win32con
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path


print = console.print


def removeDummyLaserFile(p: Path) -> None:
    if p.suffix == "" and p.stat().st_size == 0:
        try:
            os.remove(p)
        except:
            pass


def invalidNamingParts() -> None:
    laserFilePaths = util.getAllLaserFiles()
    if not laserFilePaths:
        print("All files match the naming convention!")
    invalidFilePathFoundChk = False
    for _, p in enumerate(laserFilePaths):
        if p.suffix == ".zx" or p.suffix == ".zzx":
            fileNameMatch = re.match(
                    config.RE_LASER_FILES_MATCH,
                    str(p.stem)
                    )
            if not fileNameMatch:
                invalidFilePathFoundChk = True
                print(f'------------------------\n({_}): "{p.stem}"')
    if not invalidFilePathFoundChk:
        print("没有不规范的工件名称")

def removeRedundantLaserFile() -> None:
    rawLaserFile = []

    if not config.LASER_FILE_DIR_PATH.exists():
        return

    for p in config.LASER_FILE_DIR_PATH.iterdir():
        p = util.strStandarize(p)
        if p.is_file() and "demo" not in p.stem.lower():
            rawLaserFile.append(p)

    pDeletedStr = []
    for p in rawLaserFile:
        laserFile = Path(p.parent, p.stem + ".zzx")
        if laserFile.exists() and laserFile.stat().st_mtime > p.stat().st_mtime:
            try:
                os.remove(p)
                pDeletedStr.append(str(p))
            except:
                pass

    if len(pDeletedStr) > 0:
        print(f"{len(pDeletedStr)} redundant .zx files has been deleted:")
        for pStr in pDeletedStr:
            print(pStr)
        win32api.MessageBox(
                    None,
                    f"{len(pDeletedStr)}个冗余文件已经被删除",
                    "Info",
                    4096 + 64 + 0
                )
                #   MB_SYSTEMMODAL==4096
                ##  Button Styles:
                ### 0:OK  --  1:OK|Cancel -- 2:Abort|Retry|Ignore -- 3:Yes|No|Cancel -- 4:Yes|No -- 5:Retry|No -- 6:Cancel|Try Again|Continue
                ##  To also change icon, add these values to previous number
                ### 16 Stop-sign  ### 32 Question-mark  ### 48 Exclamation-point  ### 64 Information-sign ('i' in a circle)
    else:
        print("No redundant .zx files")




def exportDimensions():
    laserFilePaths = util.getAllLaserFiles()
    wb = Workbook()
    ws = wb.create_sheet("Sheet1", 0)
    ws["A1"] = "更新时间:" + str(datetime.datetime.now().strftime("%Y-%m-%d %H%M%S%f"))
    ws.merge_cells("B1:F1")
    ws["A2"].value = "零件名称"
    ws["B2"].value = "规格"
    ws["C2"].value = "材料"
    ws["D2"].value = "第二规格指标"
    ws["E2"].value = "第二规格指标数值"
    ws["F2"].value = "长度"
    partFullNames = []
    for _, p in enumerate(laserFilePaths):
        if p.suffix == ".zx" or p.suffix == ".zzx":
            fileNameMatch = re.match(
                    config.RE_LASER_FILES_MATCH,
                    str(p.stem)
                    )
        else:
            fileNameMatch = re.match(
                    config.RE_LASER_FILES_MATCH,
                    str(p.name)
                    )

        rowMax = ws.max_row + 1

        if not fileNameMatch:
            if p.suffix == ".zx" or p.suffix == ".zzx":
                partFullName = p.stem
            else:
                partFullName = p.name
            ws[f"A{rowMax}"].value = partFullName
            ws[f"A{rowMax}"].number_format = "@"

            if partFullName in partFullNames:
                removeDummyLaserFile(p)
                continue
            else:
                partFullNames.append(partFullName)
        else:
            productId          = fileNameMatch.group(1)
            productIdNote      = fileNameMatch.group(2) # name
            partName           = fileNameMatch.group(3)
            partComponentName  = fileNameMatch.group(4) # Optional
            partMaterial       = fileNameMatch.group(5)
            partDimension               = fileNameMatch.group(6)
            part2ndDimensionInccator    = fileNameMatch.group(7) # Optional
            part2ndDimensionInccatorNum = fileNameMatch.group(9) # Optional
            partLength    = fileNameMatch.group(10)
            partDimension = partDimension.replace("_", "*")
            partDimension = partDimension.replace("x", "*")
            # partDimension = partDimension.replace("∅", "∅")
            partDimension = partDimension.replace("Ø", "∅")
            partDimension = partDimension.replace("Φ", "∅")
            partDimension = partDimension.replace("φ", "∅")
            partDimension = partDimension.strip()
            partFullName = "{} {} {}/{}".format(productId, partName, partMaterial, partDimension)
            if partFullName in partFullNames:
                removeDummyLaserFile(p)
                continue
            else:
                partFullNames.append(partFullName)

            otherPart = fileNameMatch.group(12)          # Optional
            partLongTubeLength = fileNameMatch.group(14) # Optional

            ws[f"A{rowMax}"].value = partFullName
            ws[f"A{rowMax}"].number_format = "@"
            ws[f"B{rowMax}"].value = partDimension
            ws[f"B{rowMax}"].number_format = "@"
            ws[f"C{rowMax}"].value = partMaterial
            ws[f"C{rowMax}"].number_format = "@"
            ws[f"D{rowMax}"].value = part2ndDimensionInccator
            ws[f"D{rowMax}"].number_format = "@"
            ws[f"E{rowMax}"].value = part2ndDimensionInccatorNum
            ws[f"E{rowMax}"].number_format = "0"
            ws[f"F{rowMax}"].value = partLength
            ws[f"F{rowMax}"].number_format = "0"


    # Add table
    tab = Table(displayName="Table1", ref=f"A2:F{ws.max_row}")

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
            )
    tab.tableStyleInfo = style

    ws.add_table(tab)

    util.saveWorkbook(wb, Path(config.PARENT_DIR_PATH, r"存档\零件规格总览.xlsx"), True)


