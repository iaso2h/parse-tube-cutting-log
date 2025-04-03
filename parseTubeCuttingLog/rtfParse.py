import util
import config
import console
import style

import chardet
import os
import re
import datetime
import time
from typing import Optional
from collections import Counter
from pathlib import Path
from striprtf.striprtf import rtf_to_text
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Protection
from openpyxl.styles.alignment import Alignment


def getWorkbook():
    if config.CUT_RECORD_PATH.exists():
        return load_workbook(str(config.LASER_PORFILING_PATH))
    else:
        return Workbook()
print = console.print


def getEncoding(filePath) -> str:
    # Create a magic object
    with open(filePath, "rb") as f:
        # Detect the encoding
        rawData = f.read()
        result = chardet.detect(rawData)
        if not result["encoding"]:
            return ""
        else:
            return result["encoding"]


def parse(rtfFile:Path, saveChk:bool, wb: Workbook=Workbook()) -> Optional[Workbook]:
    parseResult = {}
    laserFileOpenPat = re.compile(r"^\((.+?)\)打开文件：(.+)$")
    loopStartPat     = re.compile(r"^\((.+?)\)总零件数:(\d+), 当前零件序号:1$")
    laserFileLastOpen = ""
    loopLastTime = None
    now = datetime.datetime.now()
    with open(rtfFile, "r", encoding=getEncoding(str(rtfFile))) as f:
        content = rtf_to_text(f.read())
        lines = content.split("\n")
        for lineIdx, l in enumerate(lines):
            openMatch      = laserFileOpenPat.match(l)
            loopStartMatch = loopStartPat.match(l)
            if openMatch:
                laserFileName = openMatch.group(2)
                laserFileLastOpen = laserFileName
                if laserFileName not in parseResult:
                    parseResult[laserFileName] = {
                        "open": [],
                        "loop": [],
                        "loopIntervalUpdated": {},
                        "loopIntervalCounter": Counter(),
                        "workpieceCount": 0
                    }
                    loopLastTime = None
                parseResult[laserFileName]["open"].append(( lineIdx, openMatch.group(1) ))

            if loopStartMatch:
                timeStamp = loopStartMatch.group(1)
                timeLoop  = datetime.datetime.strptime(f"{now.year}/{timeStamp}", "%Y/%m/%d %H:%M:%S")
                if not loopLastTime:
                    loopInterval = 0
                else:
                    loopInterval = (timeLoop - loopLastTime).total_seconds()

                loopLastTime = timeLoop

                parseResult[laserFileLastOpen]["loop"].append(( lineIdx, timeStamp, loopInterval))
                parseResult[laserFileLastOpen]["loopIntervalUpdated"][f"{loopInterval}"] = loopLastTime
                parseResult[laserFileLastOpen]["loopIntervalCounter"][f"{loopInterval}"] += 1
                if parseResult[laserFileLastOpen]["workpieceCount"] == 0:
                    parseResult[laserFileLastOpen]["workpieceCount"] = loopStartMatch.group(2)

    if wb.active.title == "Sheet": # type: ignore
        ws = wb.active # type: Worksheet
        ws.title = rtfFile.stem
    else:
        ws = wb.create_sheet(rtfFile.stem, 0)

    ws[f"A{1}"].value = "排样文件"
    ws.column_dimensions["A"].width = 35
    ws[f"B{1}"].value = "循环耗时"
    ws.column_dimensions["B"].width = 12
    ws[f"C{1}"].value = "循环统计"
    ws.column_dimensions["C"].width = 12
    ws[f"D{1}"].value = "最后统计日期"
    ws.column_dimensions["D"].width = 22
    ws[f"E{1}"].value = "工件目标"
    ws.column_dimensions["E"].width = 12
    ws[f"F{1}"].value = "工件目标"
    ws.column_dimensions["F"].width = 12
    ws[f"G{1}"].value = "目标耗时"
    ws.column_dimensions["G"].width = 15
    ws[f"H{1}"].value = "预计完成时间"
    ws.column_dimensions["H"].width = 22
    for col in range(1, 8):
        ws.cell(row=1, column=col).style = "Headline 1"

    for laserFileName, laserFileInfo in parseResult.items():
        laserFileStartRow = ws.max_row
        if len(laserFileInfo["loop"]) < 1:
            continue
        mostCommon = laserFileInfo["loopIntervalCounter"].most_common(5)
        skipRowCount = 0
        for intervalIdx, common in enumerate(mostCommon):
            currentRow = intervalIdx + laserFileStartRow + 1 - skipRowCount
            interval      = common[0]
            intervalCount = common[1]

            if intervalIdx == len(mostCommon) - 1:

                # Merge laser file Name cells
                if interval == "0":
                    endRow = currentRow - 1
                else:
                    endRow = currentRow
                # Check merge necessity of merging cells
                if endRow > laserFileStartRow + 1:
                    ws.merge_cells(
                        start_row    = laserFileStartRow + 1,
                        end_row      = endRow,
                        start_column = 1,
                        end_column   = 1
                    )
                    ws.cell(row=laserFileStartRow+1,column=1).alignment = style.alCenterWrap

            # Don't fill in 0
            if interval == "0":
                skipRowCount += 1
                continue

            ws.cell(row=currentRow, column=2).value = int(float(interval))
            ws.cell(row=currentRow, column=2).number_format = '0"秒"'
            ws.cell(row=currentRow, column=3).value = intervalCount
            ws.cell(row=currentRow, column=3).number_format = '0"次"'
            ws.cell(row=currentRow, column=4).value = laserFileInfo["loopIntervalUpdated"][interval]
            ws.cell(row=currentRow, column=5).value = 100
            ws.cell(row=currentRow, column=5).number_format = '0"个"'
            ws.cell(row=currentRow, column=5).style = style.style["input"]
            ws.cell(row=currentRow, column=5).protection = Protection(locked=False)
            ws.cell(row=currentRow, column=6).value = 0
            ws.cell(row=currentRow, column=6).number_format = '0"个"'
            ws.cell(row=currentRow, column=6).style = style.style["input"]
            ws.cell(row=currentRow, column=6).protection = Protection(locked=False)
            ws.cell(row=currentRow, column=7).value = f'=(B{currentRow}+1)/{laserFileInfo["workpieceCount"]}*(E{currentRow}-F{currentRow})/86400'
            ws.cell(row=currentRow, column=7).number_format = "[h]时mm分ss秒"
            ws.cell(row=currentRow, column=8).value = f'=NOW() + G{currentRow}'
            ws.cell(row=currentRow, column=8).number_format = "yyyy-m-d h:mm:ss"


            if intervalIdx == 0:
                # Add top border
                ws[f"A{currentRow}"].border = style.borderMedium
                ws[f"B{currentRow}"].border = style.borderMedium
                ws[f"C{currentRow}"].border = style.borderMedium
                ws[f"D{currentRow}"].border = style.borderMedium
                ws[f"E{currentRow}"].border = style.borderMedium
                ws[f"F{currentRow}"].border = style.borderMedium
                ws[f"G{currentRow}"].border = style.borderMedium
                ws[f"H{currentRow}"].border = style.borderMedium
                ws.cell(row=currentRow, column=1).value = laserFileName


    ws.protection.sheet = True
    ws.protection.password = '456'
    ws.protection.enable()
    if saveChk:
        util.saveWorkbook(wb)
    else:
        return wb


def parseAllLog():
    wb = Workbook()
    for f in Path(config.LASER_LOG_PATH).iterdir():
        if f.suffix == ".rtf":
            wb = parse(f, False, wb) # type: ignore
    util.saveWorkbook(wb, config.LASER_PROFILE_PATH, True) # type: ignore


def parseWeeklyLog():
    wb = Workbook()
    now = datetime.datetime.now()
    timeDelta = datetime.timedelta(days=7)
    for f in Path(config.LASER_LOG_PATH).iterdir():
        if f.suffix == ".rtf":
            logTime = datetime.datetime.fromtimestamp(f.stat().st_ctime)
            if now - logTime <= timeDelta:
                wb = parse(f, False, wb) # type: ignore
    util.saveWorkbook(wb, config.LASER_PROFILE_PATH, True) # type: ignore
