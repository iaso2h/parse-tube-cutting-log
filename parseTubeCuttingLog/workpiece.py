import config
import console
import util
import shutil

import re
import os
import json
import datetime
import win32api, win32con
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from pathlib import Path
from openpyxl.styles.numbers import BUILTIN_FORMATS
# https://openpyxl.readthedocs.io/en/3.1.3/_modules/openpyxl/styles/numbers.html
from decimal import Decimal


print = console.print

def bankRound(precision: float, digitLiteral: str) -> float:
    return float(Decimal(digitLiteral).quantize(Decimal(precision), rounding = "ROUND_HALF_UP"))


def removeDummyLaserFile(p: Path) -> None:
    if p.suffix == "" and p.stat().st_size == 0:
        try:
            os.remove(p)
        except:
            pass


def workpieceNamingVerification() -> None:
    laserFilePaths = util.getAllLaserFiles()
    if not laserFilePaths:
        print("All files match the naming convention!")
    invalidFilePathFoundChk = False
    for _, p in enumerate(laserFilePaths):
        if p.suffix == ".zx" or p.suffix == ".zzx":
            fileNameMatch = re.match(
                    config.RE_LASER_FILES_PAT,
                    str(p.stem)
                    )
            if not fileNameMatch:
                invalidFilePathFoundChk = True
                print(f'------------------------\n({_}): "{p.stem}"')
    if not invalidFilePathFoundChk:
        print("没有不规范的工件名称")

def removeRedundantLaserFile() -> None:
    rawLaserFile = []

    if not config.LASER_FILE_DIR_PATH.exists():
        return

    for p in config.LASER_FILE_DIR_PATH.iterdir():
        p = util.strStandarize(p)
        if p.is_file() and "demo" not in p.stem.lower():
            rawLaserFile.append(p)

    pDeletedStr = []
    for p in rawLaserFile:
        laserFile = Path(p.parent, p.stem + ".zzx")
        if laserFile.exists() and laserFile.stat().st_mtime > p.stat().st_mtime:
            try:
                os.remove(p)
                pDeletedStr.append(str(p))
            except:
                pass

    if len(pDeletedStr) > 0:
        print(f"{len(pDeletedStr)} redundant .zx files has been deleted:")
        for pStr in pDeletedStr:
            print(pStr)
        win32api.MessageBox(
                    None,
                    f"{len(pDeletedStr)}个冗余文件已经被删除",
                    "Info",
                    4096 + 64 + 0
                )
                #   MB_SYSTEMMODAL==4096
                ##  Button Styles:
                ### 0:OK  --  1:OK|Cancel -- 2:Abort|Retry|Ignore -- 3:Yes|No|Cancel -- 4:Yes|No -- 5:Retry|No -- 6:Cancel|Try Again|Continue
                ##  To also change icon, add these values to previous number
                ### 16 Stop-sign  ### 32 Question-mark  ### 48 Exclamation-point  ### 64 Information-sign ('i' in a circle)
    else:
        print("No redundant .zx files")




def exportDimensions():
    laserFilePaths = util.getAllLaserFiles()
    with open(config.WORKPIECE_DICT, "r", encoding="utf-8") as f:
        workpieceDict = json.load(f)

    wb = Workbook()
    ws = wb["Sheet"]
    ws["A1"] = "更新时间:" + str(datetime.datetime.now().strftime("%Y-%m-%d %H%M%S%f"))
    ws.merge_cells("B1:F1")
    ws["A2"].value = "零件名称"
    ws.column_dimensions["A"].width = 25
    ws["B2"].value = "外发别名"
    ws.column_dimensions["B"].width = 14
    ws["C2"].value = "规格"
    ws.column_dimensions["C"].width = 20
    ws["D2"].value = "材料"
    ws.column_dimensions["D"].width = 9
    ws["E2"].value = "参数一"
    ws.column_dimensions["E"].width = 8
    ws["F2"].value = "参数二"
    ws.column_dimensions["F"].width = 8
    ws["G2"].value = "长度"
    ws.column_dimensions["G"].width = 8
    ws["H2"].value = "方数(m²)"
    ws.column_dimensions["H"].width = 9.5
    ws["I2"].value = "焊接散件"
    ws.column_dimensions["I"].width = 12
    workpieceFullNames = []
    workpieceNickNames = workpieceDict["nickname"]
    # <fullPartName>: ["<nickName>", "<comment>"]
    fileNamePat      = re.compile(config.RE_LASER_FILES_PAT)
    tubeDimensionPat = re.compile(config.TUBE_DIMENSION_PAT)
    for lIdx, p in enumerate(laserFilePaths):
        if p.suffix == ".zx" or p.suffix == ".zzx":
            fileNameMatch = fileNamePat.match(str(p.stem))
        else:
            fileNameMatch = fileNamePat.match(str(p.name))

        workpieceNickName  = ""
        workpieceDimension = ""
        surfaceAreaEval = 0

        rowMax = ws.max_row + 1

        fileNameMatchTick = False

        if not fileNameMatch:
            if p.suffix == ".zx" or p.suffix == ".zzx":
                workpieceFullName = p.stem
            else:
                workpieceFullName = p.name
            ws[f"A{rowMax}"].value = workpieceFullName
            ws[f"A{rowMax}"].number_format = "@"
            # namingly ws[f"A{rowMax}"].number_format = BUILTIN_FORMATS[49]

            if workpieceFullName.endswith(" 焊接组合"):
                workpieceNickName = workpieceFullName.replace(" 焊接组合", "")
            if workpieceFullName in workpieceNickNames:
                workpieceNickName = workpieceNickNames[workpieceFullName][0]
                if workpieceNickNames[workpieceFullName][1]:
                    comment = Comment(workpieceNickNames[workpieceFullName][1], "阮焕")
                    comment.width = 300
                    comment.height = 150
                    ws[f"B{rowMax}"].comment = comment

            ws[f"B{rowMax}"].value = workpieceNickName
            ws[f"B{rowMax}"].number_format = "@"

            if workpieceFullName in workpieceFullNames:
                removeDummyLaserFile(p)
                continue
            else:
                workpieceFullNames.append(workpieceFullName)
        else:
            fileNameMatchTick = True

            productId          = fileNameMatch.group(1)
            productIdNote      = fileNameMatch.group(2) # name
            workpieceName           = fileNameMatch.group(3)
            if workpieceName and "飞切" in workpieceName:
                workpieceName = re.sub(r"[有无]飞切", "", workpieceName)
                workpieceName = workpieceName.replace("()", "")
            workpieceComponentName  = fileNameMatch.group(4) # Optional
            workpieceMaterial       = fileNameMatch.group(5)
            workpieceDimension = fileNameMatch.group(7)
            if workpieceDimension:
                workpieceDimension = workpieceDimension.replace("_", "*")
                workpieceDimension = workpieceDimension.replace("x", "*")
                # workpieceDimension = workpieceDimension.replace("∅", "∅")
                workpieceDimension = workpieceDimension.replace("Ø", "∅")
                workpieceDimension = workpieceDimension.replace("Φ", "∅")
                workpieceDimension = workpieceDimension.replace("φ", "∅")
                workpieceDimension = workpieceDimension.strip()

            workpiece1stParameter = fileNameMatch.group(8)
            workpiece2ndParameter = fileNameMatch.group(9) # Optional

            # DEPRECATED:
            # workpiece2ndParameterNum = fileNameMatch.group(11) # Optional
            workpieceLength = fileNameMatch.group(12)

            workpieceFullName = "{} {}".format(productId, workpieceName)
            if workpieceFullName in workpieceFullNames:
                removeDummyLaserFile(p)
                continue
            else:
                workpieceFullNames.append(workpieceFullName)

            tailingWorkpiece = fileNameMatch.group(14)        # Optional
            workpieceLongTubeLength = fileNameMatch.group(16) # Optional

            ws[f"A{rowMax}"].value = workpieceFullName
            ws[f"A{rowMax}"].number_format = "@"
            # namingly ws[f"A{rowMax}"].number_format = BUILTIN_FORMATS[49]
            if workpieceFullName in workpieceNickNames:
                workpieceNickName = workpieceNickNames[workpieceFullName][0]
                comment = workpieceNickNames[workpieceFullName][1]
                if comment:
                    ws[f"B{rowMax}"].comment = Comment(comment, "阮焕")
                    ws[f"B{rowMax}"].comment.width = 300
                    ws[f"B{rowMax}"].comment.height = 150
            ws[f"B{rowMax}"].value = workpieceNickName
            ws[f"B{rowMax}"].number_format = "@"

            ws[f"C{rowMax}"].value = workpieceDimension
            ws[f"C{rowMax}"].number_format = "@"
            ws[f"D{rowMax}"].value = workpieceMaterial
            ws[f"D{rowMax}"].number_format = "@"
            ws[f"E{rowMax}"].value = workpiece1stParameter
            ws[f"E{rowMax}"].number_format = "@"
            if not workpiece2ndParameter or not re.search(r"^\d", workpiece2ndParameter):
                ws[f"F{rowMax}"].value = workpiece2ndParameter
                ws[f"F{rowMax}"].number_format = "@"
                # DEPRECATED:
                # ws[f"H{rowMax}"].value = workpiece2ndParameterNum
                # ws[f"H{rowMax}"].number_format = BUILTIN_FORMATS[2]
            ws[f"G{rowMax}"].value = workpieceLength
            ws[f"G{rowMax}"].number_format = BUILTIN_FORMATS[2]

        # Calculate the surface area
        if workpieceDimension and fileNameMatchTick and "∅" in workpieceDimension and "L" in workpieceDimension:
            m = tubeDimensionPat.match(workpieceDimension)
            if m:
                dia    = float(m.group(1)[1:])
                length = float(m.group(3)[1:])
                surfaceAreaFormula = f"=3.14 * { dia } * G{rowMax} / 1000 / 1000"
                surfaceAreaEval = 3.14 * dia * length / 1000 / 1000
                ws[f"H{rowMax}"].value = surfaceAreaFormula
                ws[f"H{rowMax}"].number_format = "0.0000"

        # Use override area
        areaOverride = workpieceDict["areaOverride"]
        if workpieceFullName in areaOverride or (workpieceNickName and workpieceNickName in areaOverride):
            if workpieceNickName:
                querryKey = workpieceNickName
            else:
                querryKey = workpieceFullName
            overrideVal = areaOverride[querryKey]

            if isinstance(overrideVal, float):
                if ws[f"H{rowMax}"].value and surfaceAreaEval:
                    print(f"Override area for {querryKey} with {areaOverride[querryKey]} instead of {surfaceAreaEval}")
                else:
                    print(f"Override area for {querryKey} with {areaOverride[querryKey]}")

                ws[f"H{rowMax}"].value = areaOverride[querryKey]
            elif isinstance(overrideVal, list):
                ws[f"I{rowMax}"].value = "\n".join(overrideVal)
                ws[f"I{rowMax}"].number_format = "@"
                ws[f"H{rowMax}"].value = f"=SUMPRODUCT(SUMIF($B:$B,TEXTSPLIT($I{rowMax},CHAR(10)),$H:$H))+SUMPRODUCT(SUMIF($A:$A,TEXTSPLIT($I{rowMax},CHAR(10)),$H:$H))"
            elif isinstance(overrideVal, str):
                ws[f"H{rowMax}"].value = f'=IF(ISNUMBER(MATCH("{overrideVal}", B:B, 0)), INDEX(H:H, MATCH("{overrideVal}", B:B, 0)), IF(ISNUMBER(MATCH("{overrideVal}", A:A, 0)), INDEX(H:H, MATCH("{overrideVal}", A:A, 0)), ""))'
                print(f"area of {querryKey} is linked to {areaOverride[querryKey]}")


            ws[f"H{rowMax}"].number_format = "0.0000"




    # Add table
    tab = Table(displayName="Table1", ref=f"A2:I{ws.max_row}")

    # Add printable area
    ws.oddFooter.center.text = "第 &[Page] 页，共 &N 页" # type: ignore
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_title_rows = "2:2"
    ws.print_area = f"A2:I{ws.max_row}"


    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
            name="TableStyleMedium16",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
            )
    tab.tableStyleInfo = style

    ws.add_table(tab)

    # Add protection
    ws.protection.sheet = True
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.password = '456'
    ws.protection.enable()

    savePath = util.saveWorkbook(wb, Path(config.PARENT_DIR_PATH, r"存档\零件规格总览.xlsx"), True)

    if os.getlogin() == "OT03":
        if config.WAREHOUSING_PATH.exists():
            shutil.copy2(savePath, Path(config.WAREHOUSING_PATH, "零件规格总览.xlsx"))



