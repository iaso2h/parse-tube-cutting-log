import util
import console
import config

import datetime
import re
import numpy
from PIL import Image, ImageFilter, ImageGrab
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from pathlib import Path


def getWorkbook() -> None:
    if config.CUT_RECORD_PATH.exists():
        return load_workbook(str(config.CUT_RECORD_PATH))
    else:
        return Workbook()

print = console.print

screenshotPaths = []
def initSheetFromScreenshots(wb: Workbook) -> None: # {{{
    yearMonthPrefix = []
    sheetNames = wb.sheetnames
    for p in config.SCREENSHOT_DIR_PATH.iterdir():
        if p.suffix == ".png":
            with Image.open(p) as img:
                width, height = img.size
                if width != 1080 or height != 1920:
                    continue

            screenshotPaths.append(p)
            dateStamp = p.stem[5:12]
            if dateStamp not in yearMonthPrefix:
                yearMonthPrefix.append(dateStamp)

    for n in yearMonthPrefix:
        if n not in sheetNames:
            ws = wb.create_sheet(n, 0)
            ws["A1"].value = "排样文件"
            ws["B1"].value = "完成时间"
            ws["C1"].value = "单号"
            ws["D1"].value = "型号(数量)"
            ws["E1"].value = "已切量/需求量"
            ws["F1"].value = "截图文件" # }}}


def takeScreenshot() -> None: # {{{
    import win32gui
    import win32process
    import psutil

    hwndTitle = {}
    def winEnumHandler(hwnd, ctx):
        if win32gui.IsWindowVisible(hwnd):
            windowText = win32gui.GetWindowText(hwnd)
            if windowText:
                hwndTitle[hwnd] = windowText
        return True

    win32gui.EnumWindows(winEnumHandler, None)

    partFileName = ""
    for hwnd, title in hwndTitle.items():
        if title.startswith("TubePro"):
            _, pId = win32process.GetWindowThreadProcessId(hwnd)
            pName = psutil.Process(pId).name()
            if pName == "TubePro.exe":
                partFileName = re.sub(r"^TubePro(\(.+?\))? (.+\.zzx).*?$", r"\2", title, re.IGNORECASE)

                win32gui.ShowWindow(hwnd, 5)
                win32gui.SetForegroundWindow(hwnd)
                break

    if not partFileName:
        return print("TubePro isn't running")

    # Check current forground program
    datetimeNow = datetime.datetime.now()
    timeStamp = datetimeNow.strftime("%Y/%m/%d %H:%M:%S")
    screenshot = ImageGrab.grab()
    screenshotPath = Path(config.SCREENSHOT_DIR_PATH, f"屏幕截图 {datetimeNow.strftime("%Y-%m-%d %H%M%S")}.png")
    screenshot.save(screenshotPath)

    # Using OCR to get process count
    wb = getWorkbook()
    sheetName = screenshotPath.stem[5:12]
    try:
        ws = wb[sheetName]
    except Exception:
        ws = wb.create_sheet(sheetName, 0)
        ws["A1"].value = "排样文件"
        ws["B1"].value = "完成时间"
        ws["C1"].value = "单号"
        ws["D1"].value = "型号(数量)"
        ws["E1"].value = "已切量/需求量"
        ws["F1"].value = "截图文件"

    newRecord(ws, screenshotPath, partFileName, timeStamp)
    util.saveWorkbook(wb, config.CUT_RECORD_PATH) # }}}
# }}}

def getImgInfo(p:Path) -> None: # {{{
    import easyocr
    import json
    reader = easyocr.Reader(["ch_sim", "en"])

    with Image.open(p) as img:
        imgTitle        = img.crop((91, 0, 900, 25))
        imgProcessCount = img.crop((550, 1665, 765, 1685))
        cvTitle = numpy.array(imgTitle)[:, :, ::-1].copy()
        cvProcessCount = numpy.array(imgProcessCount)[:, :, ::-1].copy()

        imgRGB = img.convert("RGB")
        targetCompletedPixel = imgRGB.getpixel((15, 1810))
        if targetCompletedPixel == (170, 170, 0) or targetCompletedPixel == (255, 155, 155):
            # Also treat A21 error code as completion message
            targetCompletedChk = True
        else:
            targetCompletedChk = False

        if targetCompletedChk:
            imgTimeStamp = img.crop((104, 1777, 240, 1792))
            cvTimeStamp  = numpy.array(imgTimeStamp)[:, :, ::-1].copy()
        else:
            imgTimeStamp = img.crop((91, 1755, 185, 1864)).filter(ImageFilter.EDGE_ENHANCE)
            cvTimeStamp  = numpy.array(imgTimeStamp)[:, :, ::-1].copy()

    titleRead = reader.readtext(cvTitle)
    processCountRead = reader.readtext(cvProcessCount)
    timeStampRead = reader.readtext(cvTimeStamp)
    partFileName = ""
    partProcessCount = ""
    timeStamp = p.stem[5:] # Default time stamp
    if titleRead:
        for text in titleRead:
            partFileName = partFileName + " " + text[1]
            suffixMatch = re.search(r"\.zzx", partFileName, flags=re.IGNORECASE)
            if suffixMatch:
                partFileName = partFileName[:suffixMatch.span()[1]]
            partFileName = partFileName.strip()
            with open(config.LASER_OCR_FIX_PATH, "r", encoding="utf-8") as pat:
                commonFix = json.load(pat)
            for key, val in commonFix.items():
                pattern = re.compile(key, re.IGNORECASE)
                partFileName = pattern.sub(val, partFileName)


    if processCountRead:
        if len(processCountRead) == 2:
            # In case recognition result is 2
            partProcessCount = processCountRead[1][1]


    if timeStampRead:
        timeStamp = timeStampRead[len(timeStampRead) - 1][1]
        if not targetCompletedChk:
            timeStamp = p.stem[5:9] + "/" + timeStamp # Add year prefix

        commonFix = {
                "l": "1",
                "i": "1",
                ";": ":",
                ".": ":",
                ",": ":",
                "+": ":",
                }
        for key, val in commonFix.items():
            timeStamp = timeStamp.replace(key, val)

    partFileName     = ILLEGAL_CHARACTERS_RE.sub("", partFileName)
    timeStamp        = ILLEGAL_CHARACTERS_RE.sub("", timeStamp)
    partProcessCount = ILLEGAL_CHARACTERS_RE.sub("", partProcessCount)
    return partFileName, partProcessCount, timeStamp # }}}


def validScreenshotPath(cell): # {{{
    if not cell.value or not isinstance(cell.value, str) or not Path(cell.value).exists():
        return False
    else:
        return True # }}}


def newRecord(ws, p, partFileName=None, timeStamp=None):
    if not partFileName or not timeStamp:
        partFileName, partProcessCount, timeStamp = getImgInfo(p)
    else:
        import easyocr
        reader = easyocr.Reader(["en"])
        partProcessCount = ""
        with Image.open(p) as img:
            imgProcessCount = img.crop((550, 1665, 765, 1685))
            cvProcessCount = numpy.array(imgProcessCount)[:, :, ::-1].copy()
            processCountRead = reader.readtext(cvProcessCount)
            if processCountRead:
                if len(processCountRead) == 2:
                    # In case recognition result is 2
                    partProcessCount = processCountRead[1][1]
                    partProcessCount = ILLEGAL_CHARACTERS_RE.sub("", partProcessCount)

    rowNew = ws.max_row + 1
    ws[f"A{rowNew}"].value = partFileName
    ws[f"B{rowNew}"].value = timeStamp
    ws[f"B{rowNew}"].number_format = "yyyy/m/d h:mm:ss"
    ws[f"E{rowNew}"].value = str(partProcessCount)
    ws[f"E{rowNew}"].number_format = "@"
    ws[f"F{rowNew}"].hyperlink = str(p)


def updateScreenshotRecords(): # {{{
    wb = getWorkbook()
    initSheetFromScreenshots(wb)
    for p in screenshotPaths:
        sheetName = p.stem[5:12]
        ws = wb[sheetName]
        rowMax = ws.max_row
        # fix rowMax to row that contain valid screenshot path
        lastDatetime = None
        if rowMax != 1:
            # Get the valid last datetime
            while rowMax > 1:
                lastScreenshotCell = ws[f"F{rowMax}"]
                if not validScreenshotPath(lastScreenshotCell):
                    rowMax = rowMax - 1
                    continue
                if "\n" in str(lastScreenshotCell.value).strip():
                    paths = str(lastScreenshotCell.value).strip().split("\n")
                    lastPath = Path(paths[len(paths) - 1])
                else:
                    lastPath = Path(lastScreenshotCell.value)

                try:
                    lastDatetime = datetime.datetime.strptime(str(lastPath.stem)[5:], "%Y-%m-%d %H%M%S")
                    break
                except ValueError:
                    rowMax = rowMax - 1
                    continue


            if not lastDatetime:
                newRecord(ws, p)
            else:
                currentDatetime = datetime.datetime.strptime(str(p.stem)[5:], "%Y-%m-%d %H%M%S")
                # Only save screenshots that are newer than the last one
                if lastDatetime < currentDatetime:
                    newRecord(ws, p)
        else:
            # Start in a new worksheet
            newRecord(ws, p)

    util.saveWorkbook(wb, config.CUT_RECORD_PATH) # }}}


def relinkScreenshots():
    # TODO: highlight invalid ones
    wb = getWorkbook()
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
            for cell in row:
                if not validScreenshotPath(cell):
                    continue

                if "\n" in str(cell.value).strip():
                    screenshotPaths = str(cell.value).strip().split("\n")
                    screenshotPath = Path(screenshotPaths[len(screenshotPaths) - 1])
                else:
                    screenshotPath = Path(str(cell.value))

                if screenshotPath.exists() and screenshotPath.suffix == ".png":
                    ws[f"F{cell.row}"].hyperlink = cell.value

    util.saveWorkbook(wb, config.CUT_RECORD_PATH)
