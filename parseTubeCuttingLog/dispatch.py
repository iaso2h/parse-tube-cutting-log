import config
import util
import console

import re
import copy
import json
from typing import Union, Callable
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

print = console.print
# TODO: resolve when file path not found
partColumnLetter = "E"
partColumnNum = 5
styleBorderThin = Side(border_style="thin", color="FF000000")
with open(config.PRODUCT_ID_CATERGORY_CONVENTION_PATH, "r", encoding="utf-8") as pat:
    productIdCatergoryConvention = json.load(pat)


def getRowSections(ws, colLetter: str, rowFirst: int, rowLast: int, secondSectionBreakConditionFunc: Union[None, Callable] = None,): # {{{
    sections = []

    # NOTE: ws["C"] won't yield row greater than the maximum row
    for i, cell in enumerate(ws[colLetter]):
        if sections:
            lastSectionPair = sections[len(sections) - 1]
        else:
            lastSectionPair = []

        rowNum = i + 1

        if rowNum < rowFirst:
            continue
        if rowNum == rowLast:
            if len(lastSectionPair) % 2 == 1:
                lastSectionPair.append(rowNum)
            break
        if cell.value:
            if not sections:
                sections.append([rowNum])
            else:
                if len(lastSectionPair) % 2 == 1:
                    if not secondSectionBreakConditionFunc:
                        if cell.value != ws["C{}".format(lastSectionPair[0])].value:
                            if lastSectionPair[0] != rowNum - 1:
                                lastSectionPair.append(rowNum - 1)
                                sections.append([rowNum])
                            else:
                                # Skip duplicated row, and change the 1st element of the last section pair to current row
                                lastSectionPair[0] = rowNum
                    else:
                        if secondSectionBreakConditionFunc(cell) or cell.value != ws["C{}".format(lastSectionPair[0])].value:
                            if lastSectionPair[0] != rowNum - 1:
                                lastSectionPair.append(rowNum - 1)
                                sections.append([rowNum])
                            else:
                                # Skip duplicated row, and change the 1st element of the last section pair to current row
                                lastSectionPair[0] = rowNum
                else:
                    if not secondSectionBreakConditionFunc:
                        if cell.value != ws["C{}".format(lastSectionPair[0])].value:
                            sections.append([rowNum])
                    else:
                        if secondSectionBreakConditionFunc(cell) or cell.value != ws["C{}".format(lastSectionPair[0])].value:
                            sections.append([rowNum])


    return sections # }}}


def unmergeAllCell(ws):
    rangeAllMerged = copy.copy(ws.merged_cells.ranges)
    for rng in rangeAllMerged:
        if ":" not in rng.coord:
            continue
        if "A1" not in rng and "L2" not in rng and "O1" not in rng:
            try:
                ws.unmerge_cells(rng.coord)
            except KeyError:
                pass


def unmergeCellWithin(ws, rangeAllMerged, rangeTargetTop: str, rangeTargetBot: str): # {{{
    for rng in rangeAllMerged:
        if ":" not in rng.coord:
            continue
        else:
            rangeMerged = rng.coord.split(":")

        rangeMergedTopCol = rangeMerged[0][:1]
        rangeMergedTopRow = rangeMerged[0][1:]
        rangeMergedBotCol = rangeMerged[1][:1]
        rangeMergedBotRow = rangeMerged[1][1:]
        rangeTargetTopCol = rangeTargetTop[:1]
        rangeTargetTopRow = rangeTargetTop[1:]
        rangeTargetBotCol = rangeTargetBot[:1]
        rangeTargetBotRow = rangeTargetBot[1:]
        if rangeMergedTopCol == rangeTargetTopCol and rangeMergedBotCol == rangeTargetBotCol and rangeMergedTopRow == rangeTargetTopRow and rangeMergedBotRow == rangeTargetBotRow:
            continue
        elif rangeMergedTopCol == rangeTargetTopCol and rangeMergedBotCol == rangeTargetBotCol and rangeMergedTopRow >= rangeTargetTopRow and rangeMergedBotRow <= rangeTargetBotRow:
            try:
                ws.unmerge_cells(rangeMerged[0] + ":" + rangeMerged[1]) # }}}
            except ValueError:
                pass

def fillPartInfo(): # {{{
    wb = load_workbook(str(config.DISPATCH_FILE_PATH))
    laserFilePaths = util.getAllLaserFiles()
    if not laserFilePaths:
        print(f"[red]No laser files found in: {str(config.LASER_FILE_DIR_PATH)}[/red]")
        raise SystemExit(1)


    ws = wb.active

    # Unmerge all cellss
    unmergeAllCell(ws)

    for _, p in enumerate(laserFilePaths):
        # https://regex101.com
        fileNameMatch = re.match(
                config.RE_LASER_FILES_PAT,
                str(p.stem)
                )
        if not fileNameMatch:
            continue

        productId          = fileNameMatch.group(1)
        productIdNote      = fileNameMatch.group(2) # Optional
        if not productIdNote:
            if productId in ["513L", "515L"]:
                productIdNote = "(固定式)"
            else:
                productIdNote = ""
        else:
            productId = productId.replace(productIdNote, "")

        if productId in productIdCatergoryConvention:
            productIdFullName = productIdCatergoryConvention[productId] + productIdNote + "\n" + "OT"+ productId
        else:
            productIdFullName = productIdNote + "OT" + productId

        partName           = fileNameMatch.group(3)
        partComponentName  = fileNameMatch.group(4)  # Optional
        partMaterial       = fileNameMatch.group(5)
        partDimension                  = fileNameMatch.group(6)
        partSecondDimensionInccator    = fileNameMatch.group(7) # Optional
        partSecondDimensionInccatorNum = fileNameMatch.group(9) # Optional
        partLength    = fileNameMatch.group(10)
        partDimension = partDimension.replace("_", "*")
        partDimension = partDimension.replace("x", "*")
        # partDimension = partDimension.replace("∅", "∅")
        partDimension = partDimension.replace("Ø", "∅")
        partDimension = partDimension.replace("Φ", "∅")
        partDimension = partDimension.replace("φ", "∅")
        partDimension = partDimension.strip()
        partFullName = "{} {}\n({}/{})".format(productId, partName, partMaterial, partDimension)
        otherPart = fileNameMatch.group(12)          # Optional
        partLongTubeLength = fileNameMatch.group(14) # Optional

        # DEBUG: # {{{

        # print("---------------------------------------------")
        # print(_)
        # print("Laser File:", fileNameMatch.group(0))
        # print(productIdFullName)
        # print("productId 1:", productId)
        # if productIdNote:
        #     print("productIdNote 2:", productIdNote)
        # print("partName 3:", partName)
        # if partComponentName:
        #     print("partComponentName 4:", partComponentName)
        # print("partMaterial 5:", partMaterial)
        # print("partDimension 6:", partDimension)
        # if otherPart:
        #     print("otherPart 8:", otherPart)
        # if partLongTubeLength:
        #     print("partLongTubeLength 10", partLongTubeLength)
        # print("\n")
        # }}}

        # Get product id row sections
        productIdRowSections = getRowSections(ws, "C", 4, ws.max_row, lambda cell: ws[f"B{cell.column}"].value is not None)

        existingProductId = False
        existingPartInfo = False
        for rowPair in productIdRowSections:
            coordinate = f"C{rowPair[0]}"
            if str(ws[coordinate].value).strip().replace("-5", "") == productIdFullName:
                existingProductId = True
                break

        if existingProductId:
            for rowNum in range(rowPair[0], rowPair[1] + 1):
                if not ws[f"{partColumnLetter}{rowNum}"].value:
                    continue
                if ws[f"{partColumnLetter}{rowNum}"].value and ws[f"{partColumnLetter}{rowNum}"].value == partFullName:
                    existingPartInfo = True
                    break # existing part info

            # new part
            if not existingPartInfo:
                rowNew = rowPair[1] + 1
                ws.insert_rows(rowNew)

                ws[f"{partColumnLetter}{rowNew}"].value     = partFullName
                ws[f"{partColumnLetter}{rowNew}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            rowNew = ws.max_row + 1
            ws[f"C{rowNew}"].value     = productIdFullName
            ws[f"C{rowNew}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[f"{partColumnLetter}{rowNew}"].value     = partFullName
            ws[f"{partColumnLetter}{rowNew}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    util.saveWorkbook(wb, config.DISPATCH_FILE_PATH) # }}}


def beautifyCells(): # {{{
    wb = load_workbook(str(config.DISPATCH_FILE_PATH))
    ws = wb.active
    rowMax = ws.max_row
    colMax = ws.max_column
    rangeAllMerged = copy.copy(ws.merged_cells.ranges)
    # Get product id row sections
    productIdRowSections = getRowSections(ws, "C", 4, rowMax, lambda cell: ws[f"B{cell.column}"].value is not None)


    # Merge product Id rows
    for rowPair in productIdRowSections:
        unmergeCellWithin(ws, rangeAllMerged, f"C{rowPair[0]}", f"C{rowPair[1]}")
        ws.merge_cells(f"C{rowPair[0]}:C{rowPair[1]}")


    # Fill sequence number
    for rowPair in productIdRowSections:
        for i, rowNum in enumerate(range(rowPair[0], rowPair[1] + 1)):
            sequenceNum = i + 1
            ws[f"A{rowNum}"].value = sequenceNum


    # Merge order sequence rows
    for rowPair in productIdRowSections:
        orderSequence = ""
        for rowOrder in ws.iter_rows(min_col=2, max_col=2, min_row=rowPair[0], max_row=rowPair[1]):
            if orderSequence:
                unmergeCellWithin(ws, rangeAllMerged, f"B{rowPair[0]}", f"B{rowPair[1]}")
                ws.merge_cells(f"B{rowPair[0]}:B{rowPair[1]}")
                ws[f"B{rowPair[0]}"].value = orderSequence

            for cellOrder in rowOrder:
                if cellOrder.value:
                    orderSequence = cellOrder.value
                    break

    # Merge order number rows
    for rowPair in productIdRowSections:
        orderNum = ""
        for rowOrder in ws.iter_rows(min_col=4, max_col=4, min_row=rowPair[0], max_row=rowPair[1]):
            if orderNum:
                unmergeCellWithin(ws, rangeAllMerged, f"D{rowPair[0]}", f"D{rowPair[1]}")
                ws.merge_cells(f"D{rowPair[0]}:D{rowPair[1]}")
                ws[f"D{rowPair[0]}"].value = orderNum
                break

            for cellOrder in rowOrder:
                if cellOrder.value:
                    orderNum = cellOrder.value
                    break


    # Merge part info rows
    # Get part info row sections
    for rowPair in productIdRowSections:
        partInfoRowSections = getRowSections(ws, "E", rowPair[0], rowPair[1])
        for rowPartPair in partInfoRowSections:
            unmergeCellWithin(ws, rangeAllMerged, f"E{rowPartPair[0]}", f"E{rowPartPair[1]}")
            ws.merge_cells(f"E{rowPartPair[0]}:E{rowPartPair[1]}")

    # Add border to cells
    for row in ws[f"A3:P{rowMax}"]:
        for cell in row:
            cell.border = Border(
                    top=styleBorderThin,
                    left=styleBorderThin,
                    right=styleBorderThin,
                    bottom=styleBorderThin
                    )
            if cell.coordinate[0] in ["A", "B", "D"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif cell.coordinate[0] in ["C", "E"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add filter button
    fullRange = "A3:" + get_column_letter(ws.max_column)  + str(ws.max_row)
    ws.auto_filter.ref = fullRange

    util.saveWorkbook(wb, config.DISPATCH_FILE_PATH) # }}}
